"""
Build 27S Universities sheet template (xlsx) with:
- 5 sheets: master_universities, americas, asia_oceania, europe, notes
- Headers with comments (description per column)
- Sample rows in red background
- Data validation dropdowns on enum columns
- Frozen first row, reasonable column widths
"""
from pathlib import Path

from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter

OUT_PATH = r"D:\projects\oia\20-univ-finder\2027S\27S_universities_template.xlsx"

# --------------------------------------------------------------------------
# Column definitions: (key, header_comment, sample_americas, sample_asia, sample_europe)
# --------------------------------------------------------------------------
MAIN_COLUMNS = [
    # A. Basic (7)
    ("university_name",
     "공식 영문 학교명. 단과대 단독 프로그램은 'XYZ University - College of Y' 형식.\n예: Harvard University",
     "University of Washington", "The University of Tokyo", "Sciences Po"),
    ("website",
     "학교 공식 메인 URL.\n예: https://www.harvard.edu",
     "https://www.washington.edu", "https://www.u-tokyo.ac.jp/en/", "https://www.sciencespo.fr/en"),
    ("continent",
     "Americas / Asia/Oceania / Europe / Other 중 하나. 드롭다운 선택.",
     "Americas", "Asia/Oceania", "Europe"),
    ("country",
     "공식 영문 국가명.\n예: United States, France, Japan",
     "United States", "Japan", "France"),
    ("language_of_instruction_1",
     "English 또는 빈칸 두 값만. 영어 수업 제공하면 English, 아니면 빈칸.",
     "English", "English", ""),
    ("language_of_instruction_2",
     "French / German / Spanish / Russian / Japanese / Chinese / 빈칸 중 하나. 비영어 수업 언어.",
     "", "Japanese", "French"),
    ("program_type",
     "Regular / UIC Exclusive / ISEP 등. 일반 교환은 Regular.",
     "Regular", "Regular", "Regular"),

    # B. Quota & eligibility (5)
    ("quota",
     "이번 학기 모집 인원 수 (숫자).\n예: 2",
     2, 3, 2),
    ("quota_unit",
     "slot / person 중 하나. slot은 학기별 자리 단위, person은 인원 단위.",
     "slot", "person", "slot"),
    ("accepts_undergrad",
     "Y / N 두 값만. 학부생 모집 여부.",
     "Y", "Y", "Y"),
    ("accepts_grad",
     "Y / N 두 값만. 대학원생 모집 여부.",
     "Y", "Y", "N"),
    ("gpa_required",
     "4.3 만점 기준 최소 GPA (숫자).\n예: 3.0",
     3.0, 3.0, 3.3),

    # C. Semester options (5)
    ("fall_one_semester",
     "Y / N. Fall 한 학기만 가능 여부.",
     "Y", "Y", "Y"),
    ("fall_calendar_year",
     "Y / N. Fall부터 시작해 1년(Fall+Spring) 가능 여부.",
     "Y", "N", "Y"),
    ("spring_one_semester",
     "Y / N. Spring 한 학기만 가능 여부.",
     "Y", "Y", "Y"),
    ("spring_calendar_year",
     "Y / N. Spring부터 시작해 1년(Spring+다음 Fall) 가능 여부.",
     "N", "N", "N"),
    ("academic_system",
     "Semester / Quarter / Trimester / Other 중 하나.",
     "Quarter", "Semester", "Semester"),

    # D. English requirements (6)
    ("toefl_total",
     "TOEFL iBT 총점 (0~120). 요구 없으면 빈칸.\n예: 90",
     92, 80, ""),
    ("toefl_reading",
     "TOEFL Reading 서브스코어 (선택).",
     22, 20, ""),
    ("toefl_listening",
     "TOEFL Listening 서브스코어 (선택).",
     22, 20, ""),
    ("toefl_speaking",
     "TOEFL Speaking 서브스코어 (선택).",
     22, 20, ""),
    ("toefl_writing",
     "TOEFL Writing 서브스코어 (선택).",
     22, 20, ""),
    ("english_notes",
     "추가 설명.\n예: 'or equivalent, TOEFL subscore 4.0'",
     "subscores must each be at least 20", "or equivalent", ""),

    # E. Non-English requirements (2)
    ("lang2_level",
     "A1~C2 (CEFR: French/German/Spanish/Russian),\n"
     "N5~N1 (JLPT: Japanese),\n"
     "HSK1~HSK6 (HSK: Chinese) 중 하나.\n"
     "비영어 수업 없으면 빈칸.",
     "", "N2", "B2"),
    ("language_other_notes",
     "어학 요건 추가 메모.",
     "", "JLPT N2 required for non-English track", "B2 required, C1 recommended"),

    # F. Major/admission (3)
    ("available_areas",
     "수강 가능 전공/학과 안내.",
     "All majors except Medicine, Dentistry",
     "All faculties open to exchange",
     "Limited to Social Sciences and Humanities"),
    ("restricted_areas",
     "수강 불가 전공/학과.\n예: Medicine, MBA",
     "Medicine, Dentistry",
     "Medicine, Law graduate",
     "Sciences Po Law School"),
    ("admission_notes",
     "기타 입학·심사 시 유의사항.",
     "Strong academic record required.",
     "Statement of purpose required.",
     "Motivation letter in French preferred."),

    # G. Dates (5)
    ("fall_nomination_deadline",
     "MM/DD 또는 'Apr 15' 형식.",
     "04/01", "04/15", "04/10"),
    ("fall_application_deadline",
     "MM/DD 또는 'May 15' 형식.",
     "05/01", "05/15", "05/10"),
    ("spring_nomination_deadline",
     "MM/DD 형식.",
     "10/01", "10/15", "10/10"),
    ("spring_application_deadline",
     "MM/DD 형식.",
     "10/15", "10/30", "10/25"),
    ("academic_calendar",
     "학사 일정 URL 또는 학기 시작-종료일 텍스트.",
     "https://www.washington.edu/students/reg/calendar.html",
     "https://www.u-tokyo.ac.jp/en/students/academic-calendar.html",
     "Sep - Jun"),

    # H. Additional (4)
    ("housing_info",
     "기숙사 안내 URL 또는 설명.",
     "https://hfs.uw.edu/",
     "On-campus housing available",
     "https://www.sciencespo.fr/housing"),
    ("housing_guaranteed",
     "Yes / No / Partial / Unknown 중 하나.",
     "No", "Partial", "No"),
    ("recent_factsheet",
     "최근 학기 factsheet PDF URL.",
     "https://example.com/uw_factsheet.pdf",
     "https://example.com/utokyo_factsheet.pdf",
     "https://example.com/sciencespo_factsheet.pdf"),
    ("additional_link",
     "기타 홍보·안내 링크.",
     "", "", ""),
]

NOTES_COLUMNS = [
    ("university_name",
     "master_universities의 university_name과 일치해야 함 (FK).",
     "Sciences Po"),
    ("labels",
     "Scholarship / New / Popular / Update 중 해당되는 것을 쉼표로 구분.\n예: Scholarship, Popular",
     "Scholarship, Update"),
    ("note_text",
     "메모 본문. 학생에게 노출되는 안내문.",
     "Erasmus+ scholarship available for selected students."),
    ("note_link",
     "메모 관련 외부 링크 (선택).",
     "https://www.sciencespo.fr/scholarships"),
    ("update_date",
     "YYYY-MM-DD 형식.",
     "2027-01-15"),
    ("update_description",
     "이번 학기 갱신 사유. 학생/팀이 변경점을 확인할 수 있도록 한 줄로.",
     "Quota reduced from 3 to 2; new scholarship added."),
]

# Content sheet — 학교별 상세 페이지용 (master_universities와 join)
# Sorbonne 데모 row는 sorbonne_seed.csv에서 자동으로 읽어와 채워짐
CONTENT_COLUMNS = [
    ("university_name",
     "master_universities의 university_name과 일치해야 함 (FK).\n예: Sorbonne University - Faculty of Letters"),
    ("cover_image_url",
     "학교 대표 사진 URL (외부 호스팅 1개). 비어있으면 placeholder 표시.\n예: https://www.school.edu/campus.jpg"),
    ("cover_image_credit",
     "대표 사진 출처/저작권 표시. 비어있으면 표시 안 함.\n예: © Sorbonne Université / 학교 공식"),
    ("gallery_urls",
     "추가 사진 URL. 한 줄에 1개씩 (Shift+Enter로 줄바꿈).\n비어있으면 갤러리 섹션 비표시. 최대 6장.\n예:\nhttps://url1.jpg\nhttps://url2.jpg"),
    ("video_links",
     "동영상 URL. 한 줄에 1개. 형식: 라벨 | URL\n비어있으면 동영상 섹션 비표시. 최대 3개.\n예:\nCampus Tour | https://youtu.be/abc\nWelcome message | https://youtu.be/xyz"),
    ("about_text",
     "학교 소개 2-3문단. 줄바꿈 = Shift+Enter.\n간단한 마크다운 가능: **굵게**, [텍스트](url)"),
    ("why_yonsei",
     "왜 이 학교가 연세대 학생에게 좋은지 OIA가 정제해서 작성.\n간단한 마크다운 가능."),
    ("campus_location",
     "캠퍼스 위치·도시 분위기·교통. student_tips의 캠퍼스/주변 환경 인용을 정제해 옮기는 것을 권장."),
    ("accommodation",
     "주거 안내. 학교 제공 옵션, 기숙사, 외부 플랫폼 등."),
    ("student_tips",
     "선별·정제된 학생 팁. 한 줄 = 1 tip 권장.\n예:\n- 수강신청 복잡, 학과 coordinator와 메일로 진행\n- 한국관 기숙사 추천 (Cité Universitaire)"),
    ("useful_links",
     "유용한 외부 링크. 한 줄에 1개. 형식: 라벨 | URL\n비어있으면 섹션 비표시. 최대 5개.\n예:\nCourse catalogue | https://...\nVirtual tour | https://..."),
    ("last_updated",
     "이 학교 콘텐츠 마지막 갱신일. YYYY-MM-DD 형식.\n예: 2026-05-20"),
]

# --------------------------------------------------------------------------
# Styles
# --------------------------------------------------------------------------
HEADER_FILL = PatternFill(start_color="003876", end_color="003876", fill_type="solid")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
SAMPLE_FILL = PatternFill(start_color="FFE5E5", end_color="FFE5E5", fill_type="solid")
SAMPLE_FONT = Font(name="Calibri", size=10, italic=True, color="990000")
THIN_BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)

# Enum dropdown values
ENUM_VALUES = {
    "continent": ["Americas", "Asia/Oceania", "Europe", "Other"],
    "language_of_instruction_1": ["English"],
    "language_of_instruction_2": ["French", "German", "Spanish", "Russian", "Japanese", "Chinese"],
    "program_type": ["Regular", "UIC Exclusive", "ISEP", "Other"],
    "quota_unit": ["slot", "person"],
    "accepts_undergrad": ["Y", "N"],
    "accepts_grad": ["Y", "N"],
    "fall_one_semester": ["Y", "N"],
    "fall_calendar_year": ["Y", "N"],
    "spring_one_semester": ["Y", "N"],
    "spring_calendar_year": ["Y", "N"],
    "academic_system": ["Semester", "Quarter", "Trimester", "Other"],
    "lang2_level": [
        "A1", "A2", "B1", "B2", "C1", "C2",
        "N5", "N4", "N3", "N2", "N1",
        "HSK1", "HSK2", "HSK3", "HSK4", "HSK5", "HSK6",
    ],
    "housing_guaranteed": ["Yes", "No", "Partial", "Unknown"],
}

NOTES_ENUM_VALUES = {
    # labels can be multi-value comma-separated, so no strict dropdown for labels.
    # We add a free-text but mention allowed tokens in comment.
}

# Column widths (tweak for readability)
COL_WIDTH_DEFAULT = 18
COL_WIDTH_OVERRIDES = {
    "university_name": 32,
    "website": 36,
    "country": 18,
    "english_notes": 32,
    "language_other_notes": 28,
    "available_areas": 36,
    "restricted_areas": 28,
    "admission_notes": 36,
    "academic_calendar": 36,
    "housing_info": 32,
    "recent_factsheet": 32,
    "additional_link": 28,
    "note_text": 50,
    "note_link": 32,
    "update_description": 40,
}

# Wider widths for the content sheet (long-form text)
CONTENT_COL_WIDTH = {
    "university_name":    32,
    "cover_image_url":    32,
    "cover_image_credit": 24,
    "gallery_urls":       28,
    "video_links":        36,
    "about_text":         50,
    "why_yonsei":         40,
    "campus_location":    40,
    "accommodation":      40,
    "student_tips":       50,
    "useful_links":       40,
    "last_updated":       14,
}


def build_sheet(wb, sheet_name, columns, sample_index=None):
    """
    Build one sheet. `columns` is list of tuples (key, comment, *samples).
    `sample_index` picks which sample to populate (0=americas, 1=asia, 2=europe).
    If None, populates ALL three samples (used for master_universities).
    """
    ws = wb.create_sheet(sheet_name)

    # Header row
    for col_idx, col in enumerate(columns, start=1):
        key, comment_text = col[0], col[1]
        cell = ws.cell(row=1, column=col_idx, value=key)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
        cell.comment = Comment(comment_text, "OIA")
        cell.comment.width = 320
        cell.comment.height = 140

        letter = get_column_letter(col_idx)
        ws.column_dimensions[letter].width = COL_WIDTH_OVERRIDES.get(key, COL_WIDTH_DEFAULT)

    ws.row_dimensions[1].height = 32

    # Sample rows
    if sample_index is None:
        # master_universities: include 3 samples (americas, asia, europe)
        sample_rows = [2, 3, 4]
        sample_indices = [0, 1, 2]
    else:
        sample_rows = [2]
        sample_indices = [sample_index]

    for row_offset, s_idx in zip(sample_rows, sample_indices):
        for col_idx, col in enumerate(columns, start=1):
            value = col[2 + s_idx] if len(col) > 2 + s_idx else ""
            cell = ws.cell(row=row_offset, column=col_idx, value=value)
            cell.fill = SAMPLE_FILL
            cell.font = SAMPLE_FONT
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            cell.border = THIN_BORDER

    # Freeze header
    ws.freeze_panes = "A2"

    # Data validation dropdowns
    for col_idx, col in enumerate(columns, start=1):
        key = col[0]
        if key in ENUM_VALUES:
            values = ENUM_VALUES[key]
            formula = '"' + ",".join(values) + '"'
            dv = DataValidation(type="list", formula1=formula, allow_blank=True)
            dv.error = f"허용되는 값만 입력하세요: {', '.join(values)}"
            dv.errorTitle = "잘못된 값"
            dv.prompt = f"드롭다운에서 선택: {', '.join(values)}"
            dv.promptTitle = key
            letter = get_column_letter(col_idx)
            dv.add(f"{letter}2:{letter}1000")
            ws.add_data_validation(dv)

    return ws


def build_content_sheet(wb, seed_csv_path=None):
    """Build the `content` sheet (학교별 상세 페이지용).

    If seed_csv_path is provided and exists, the first data row is auto-filled
    from that CSV (typically the output of _build_content_seed.py)."""
    import csv as csvmod

    ws = wb.create_sheet("content")

    # Header row
    for col_idx, (key, comment) in enumerate(CONTENT_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=key)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
        cell.comment = Comment(comment, "OIA")
        cell.comment.width = 380
        cell.comment.height = 180
        letter = get_column_letter(col_idx)
        ws.column_dimensions[letter].width = CONTENT_COL_WIDTH.get(key, 30)
    ws.row_dimensions[1].height = 40

    # Seed Sorbonne row (or whatever is in the seed CSV) as a sample row
    if seed_csv_path and Path(seed_csv_path).exists():
        with open(seed_csv_path, encoding="utf-8-sig") as f:
            reader = csvmod.DictReader(f)
            for row_idx, row in enumerate(reader, start=2):
                for col_idx, (key, _) in enumerate(CONTENT_COLUMNS, start=1):
                    val = row.get(key, "")
                    cell = ws.cell(row=row_idx, column=col_idx, value=val)
                    cell.fill = SAMPLE_FILL
                    cell.font = SAMPLE_FONT
                    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                    cell.border = THIN_BORDER
                # Tall row to accommodate long-form text preview
                ws.row_dimensions[row_idx].height = 240

    # Freeze header row + university_name column
    ws.freeze_panes = "B2"
    return ws


def build_workbook():
    wb = Workbook()
    wb.remove(wb.active)

    # 1) master_universities — 3 samples (americas + asia + europe)
    build_sheet(wb, "master_universities", MAIN_COLUMNS, sample_index=None)

    # 2-4) Region sheets — 1 sample each
    build_sheet(wb, "americas", MAIN_COLUMNS, sample_index=0)
    build_sheet(wb, "asia_oceania", MAIN_COLUMNS, sample_index=1)
    build_sheet(wb, "europe", MAIN_COLUMNS, sample_index=2)

    # 5) notes
    build_sheet(wb, "notes", NOTES_COLUMNS, sample_index=0)

    # 6) content — 학교별 상세 페이지. Sorbonne seed CSV가 있으면 자동 포함
    seed_csv = Path(__file__).parent / "sorbonne_seed.csv"
    build_content_sheet(wb, seed_csv if seed_csv.exists() else None)

    wb.save(OUT_PATH)
    print(f"Saved: {OUT_PATH}")


if __name__ == "__main__":
    build_workbook()
